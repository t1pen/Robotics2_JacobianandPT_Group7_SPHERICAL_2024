from pathlib import Path
from tkinter import *
import openpyxl
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, messagebox
import roboticstoolbox as rtb
import numpy as np 
from roboticstoolbox import DHRobot, RevoluteDH, PrismaticDH
from datetime import datetime
import os 
import tkinter as tk
import pyglet
import sympy as sp

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH1 = OUTPUT_PATH / Path(r"assets")
ASSETS_PATH2 = OUTPUT_PATH / Path(r"assets\Jacobian")

def relative_to_assets1(path: str) -> Path:
    return ASSETS_PATH1 / Path(path)

def relative_to_assets2(path: str) -> Path:
    return ASSETS_PATH2 / Path(path)

# Load External Fonts
pyglet.options['win32_gdi_font'] = True
pyglet.font.add_file(r'assets\metropolis.extra-bold.otf')
pyglet.font.add_file(r'assets\metropolis.semi-bold.otf')
pyglet.font.add_file(r'assets\metropolis.black.otf')

window = Tk()
# Make Window Appear at the middle of the screen
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
window_width = 900
window_height = 600
x_position = (screen_width - window_width) // 2
y_position = (screen_height - window_height) // 2
window.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

# Configure Window Title and Icon
window.configure(bg = "#FFFFFF")
window.title("Spherical Manipulator Calculator")
window.iconbitmap(r"assets\robotic-arm.ico")

# Create a workbook and add sheets for f_k and i_k data
workbook = openpyxl.Workbook()
f_k_sheet = workbook.create_sheet(title="f_k_data")
i_k_sheet = workbook.create_sheet(title="i_k_data")

# Counter variables for row indices in each sheet
f_k_row_index = 1
i_k_row_index = 1

# Define a Boolean variable to track the state of the checkbutton
save_to_excel_var = BooleanVar()
save_to_excel_var.set(True)  # Set to off by default

J_sw = None


def toggle_save_to_excel():
    global save_to_excel_var

def reset():
    for entry in [a1_E, a2_E, a3_E, t1_E, t2_E, d3_E, x_E, y_E, z_E]:
        entry.delete(0, END)


def save_to_excel_fk():
    global workbook, f_k_row_index, save_to_excel_var

    # Check if the save to Excel checkbox is checked
    if not save_to_excel_var.get():
        messagebox.showinfo("Info", "Data will not be saved to Excel.")
        return

    # Check if the workbook exists
    if Path("spherical_manipulator_data.xlsx").exists():
        workbook = openpyxl.load_workbook("spherical_manipulator_data.xlsx")
    else:
        workbook = openpyxl.Workbook()

    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Get values from entry widgets
    a1_val = a1_E.get()
    a2_val = a2_E.get()
    a3_val = a3_E.get()
    t1_val = t1_E.get()
    t2_val = t2_E.get()
    d3_val = d3_E.get()
    x_val = x_E.get()
    y_val = y_E.get()
    z_val = z_E.get()

    # Get or create sheets for f_k and i_k data
    f_k_sheet = workbook["f_k_data"] if "f_k_data" in workbook.sheetnames else workbook.create_sheet("f_k_data")

    # Delete the first sheet if it's empty
    if len(workbook.sheetnames) > 1 and workbook.sheetnames[0] == 'Sheet':
        del workbook['Sheet']

    if f_k_sheet.max_row == 1:
        f_k_sheet.append(["Timestamp", "a1", "a2", "a3", "t1", "t2", "d3", "x", "y", "z"])


    # Save data to respective sheets
    f_k_sheet.append([timestamp, a1_val, a2_val, a3_val, t1_val, t2_val, d3_val, x_val, y_val, z_val])

    # Increment row indices for next data entry
    f_k_row_index += 1

    # Save the workbook
    workbook.save("spherical_manipulator_data.xlsx")

    # Notify user that data has been saved
    messagebox.showinfo("Data Saved", "Data has been saved to spherical_manipulator_data.xlsx")

def save_to_excel_ik():
    global workbook, i_k_row_index, save_to_excel_var

    # Check if the save to Excel checkbox is checked
    if not save_to_excel_var.get():
        messagebox.showinfo("Info", "Data will not be saved to Excel.")
        return

    # Check if the workbook exists
    if Path("spherical_manipulator_data.xlsx").exists():
        workbook = openpyxl.load_workbook("spherical_manipulator_data.xlsx")
    else:
        workbook = openpyxl.Workbook()

    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Get values from entry widgets
    a1_val = a1_E.get()
    a2_val = a2_E.get()
    a3_val = a3_E.get()
    t1_val = t1_E.get()
    t2_val = t2_E.get()
    d3_val = d3_E.get()
    x_val = x_E.get()
    y_val = y_E.get()
    z_val = z_E.get()

    # Get or create sheets for i_k data
    i_k_sheet = workbook["i_k_data"] if "i_k_data" in workbook.sheetnames else workbook.create_sheet("i_k_data")

    # Delete the first sheet if it's empty
    if len(workbook.sheetnames) > 1 and workbook.sheetnames[0] == 'Sheet':
        del workbook['Sheet']

    if i_k_sheet.max_row == 1:
        i_k_sheet.append(["Timestamp", "a1", "a2", "a3", "x", "y", "z", "t1", "t2", "d3"])

    # Save data to respective sheets
    i_k_sheet.append([timestamp, a1_val, a2_val, a3_val, x_val, y_val, z_val, t1_val, t2_val, d3_val])

    # Increment row indices for next data entry
    i_k_row_index += 1

    # Save the workbook
    workbook.save("spherical_manipulator_data.xlsx")

    # Notify user that data has been saved
    messagebox.showinfo("Data Saved", "Data has been saved to spherical_manipulator_data.xlsx")

def view_excel_file():
    try:
        os.startfile("spherical_manipulator_data.xlsx")
    except Exception as e:
        messagebox.showerror("Error", f"Error opening Excel file: {e}")

def close_J_sw():
    if 'J_sw' in globals() and J_sw is not None:
        J_sw.destroy()

def p_t_pick():

    ## Create Spherical Manipulator Model
    # link lengths in cm
    a1 = float(50)
    a2 = float(20)
    a3 = float(15)

    #link conversion to meters
    def cm_to_meter(a):
        m = 100 # 1 meter  = 100 cm
        return a/m

    def deg2rad(T):
        return (T/180)*np.pi

    a1 = cm_to_meter(a1)
    a2 = cm_to_meter(a2)
    a3 = cm_to_meter(a3)

    # limit of variable d1
    d3 = float(25)
    d3 = cm_to_meter(d3)

    #Create links
    #robot_variable = DHRobot([RevoluteDH(d,r,alpha,offset=theta,qlim)])
    #robot_variable = DHRobot([PrismaticDH(d=0,r,alpha,offset=d,qlim)])
    Spherical = DHRobot([
    RevoluteDH(a1,0,(90.0/180.0)*np.pi,(0.0/180.0)*np.pi,qlim=[deg2rad(-180),deg2rad(180)]),
    RevoluteDH(0,0,(90.0/180.0)*np.pi,(90.0/180.0)*np.pi,qlim=[-np.pi/4,np.pi/4]),
    PrismaticDH(0,0,(0.0/180)*np.pi,a2+a3,qlim=[0,d3])
    ], name="Spherical")

    print(Spherical)

    ## Path Trajectory Planning 

    # Q Paths (Joint Paths)
    # Spherical Joint Variables = ([T1, T2, d3])

    # Degrees to Radian


    q0 = np.array([0,0,0]) # origin

    q1 = np.array([deg2rad(float(45)),
            deg2rad(float(-20)),
            cm_to_meter(float(0))
            ])

    q2 = np.array([deg2rad(float(45)),
            deg2rad(float(-20)),
            cm_to_meter(float(25))
            ])

    q3 = np.array([deg2rad(float(45)),
            deg2rad(float(0)),
            cm_to_meter(float(0))
            ])

    q4 = np.array([deg2rad(float(-180)),
            deg2rad(float(-20)),
            cm_to_meter(float(0))
            ])
            
    q5 = np.array([deg2rad(float(-180)),
            deg2rad(float(-20)),
            cm_to_meter(float(25))
            ])

    q6 = np.array([deg2rad(float(45)),
            deg2rad(float(0)),
            cm_to_meter(float(0))
            ])

    q7 = np.array([deg2rad(float(45)),
            deg2rad(float(-20)),
            cm_to_meter(float(0))
            ])

    q8 = np.array([deg2rad(float(45)),
            deg2rad(float(-20)),
            cm_to_meter(float(25))
            ])

    q9 = np.array([deg2rad(float(45)),
            deg2rad(float(0)),
            cm_to_meter(float(0))
            ])

    q10 = np.array([deg2rad(float(180)),
            deg2rad(float(-20)),
            cm_to_meter(float(0))
            ])

    q11 = np.array([deg2rad(float(180)),
            deg2rad(float(-20)),
            cm_to_meter(float(25))
            ])

    #Plot Scale
    x1 = -1.0
    x2 = 1.0
    y1 = -1.0
    y2 = 1.0
    z1 = -0.1
    z2 = 1.0

    # Trajectory Formula
    traj1 = rtb.jtraj(q0,q1,5)
    traj2 = rtb.jtraj(q1,q2,5)
    traj3 = rtb.jtraj(q2,q3,10)
    traj4 = rtb.jtraj(q3,q4,10)
    traj5 = rtb.jtraj(q4,q5,10)
    traj6 = rtb.jtraj(q5,q6,5)
    traj7 = rtb.jtraj(q6,q7,5)
    traj8 = rtb.jtraj(q7,q8,5)
    traj9 = rtb.jtraj(q8,q9,10)
    traj10 = rtb.jtraj(q9,q10,10)
    traj11 = rtb.jtraj(q10,q11,10)
    traj12 = rtb.jtraj(q11,q0,5)


    # plot formula
    Spherical.plot(traj1.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj2.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj3.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj4.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj5.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj6.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj7.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj8.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj9.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj10.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj11.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj12.q, limits=[x1,x2,y1,y2,z1,z2], block=True)

def p_t_weld():
    ## Create Spherical Manipulator Model
    # link lengths in cm
    a1 = float(50)
    a2 = float(20)
    a3 = float(15)

    #link conversion to meters
    def cm_to_meter(a):
        m = 100 # 1 meter  = 100 cm
        return a/m

    a1 = cm_to_meter(a1)
    a2 = cm_to_meter(a2)
    a3 = cm_to_meter(a3)

    # limit of variable d1
    d3 = float(25)
    d3 = cm_to_meter(d3)

    #Create links
    #robot_variable = DHRobot([RevoluteDH(d,r,alpha,offset=theta,qlim)])
    #robot_variable = DHRobot([PrismaticDH(d=0,r,alpha,offset=d,qlim)])
    Spherical = DHRobot([
        RevoluteDH(a1,0,(90.0/180.0)*np.pi,(0.0/180.0)*np.pi,qlim=[-np.pi/2,np.pi/2]),
        RevoluteDH(0,0,(90.0/180.0)*np.pi,(90.0/180.0)*np.pi,qlim=[-np.pi/4,np.pi/4]),
        PrismaticDH(0,0,(0.0/180)*np.pi,a2+a3,qlim=[0,d3])
    ], name="Spherical")

    print(Spherical)

    ## Path Trajectory Planning 

    # Q Paths (Joint Paths)
    # Spherical Joint Variables = ([T1, T2, d3])

    # Degrees to Radian
    def deg2rad(T):
        return (T/180)*np.pi

    q0 = np.array([0,0,0]) # origin

    q1 = np.array([deg2rad(float(-45)),
                deg2rad(float(-45)),
                cm_to_meter(float(0))
                ])

    q2 = np.array([deg2rad(float(-45)),
                deg2rad(float(-45)),
                cm_to_meter(float(25))
                ])

    q3 = np.array([deg2rad(float(-45)),
                deg2rad(float(45)),
                cm_to_meter(float(0))
                ])

    q4 = np.array([deg2rad(float(-45)),
                deg2rad(float(45)),
                cm_to_meter(float(25))
                ])
                
    q5 = np.array([deg2rad(float(45)),
                deg2rad(float(45)),
                cm_to_meter(float(0))
                ])

    q6 = np.array([deg2rad(float(45)),
                deg2rad(float(45)),
                cm_to_meter(float(25))
                ])

    q7 = np.array([deg2rad(float(45)),
                deg2rad(float(-45)),
                cm_to_meter(float(0))
                ])

    q8 = np.array([deg2rad(float(45)),
                deg2rad(float(-45)),
                cm_to_meter(float(25))
                ])

    #Plot Scale
    x1 = -0.2
    x2 = 1.0
    y1 = -1.0
    y2 = 1.0
    z1 = -0.1
    z2 = 1.0

    # Trajectory Formula
    traj1 = rtb.jtraj(q0,q1,7)
    traj2 = rtb.jtraj(q1,q2,5)
    traj3 = rtb.jtraj(q2,q3,7)
    traj4 = rtb.jtraj(q3,q4,5)
    traj5 = rtb.jtraj(q4,q5,7)
    traj6 = rtb.jtraj(q5,q6,5)
    traj7 = rtb.jtraj(q6,q7,7)
    traj8 = rtb.jtraj(q7,q8,5)
    traj9 = rtb.jtraj(q8,q0,7)


    # plot formula
    Spherical.plot(traj1.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj2.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj3.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj4.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj5.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj6.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj7.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj8.q, limits=[x1,x2,y1,y2,z1,z2])
    Spherical.plot(traj9.q, limits=[x1,x2,y1,y2,z1,z2], block=True)

def f_k():

    close_J_sw()
    
 # Get values from entry widgets
    a1 = a1_E.get().strip()
    a2 = a2_E.get().strip()
    a3 = a3_E.get().strip()
    t1 = t1_E.get().strip()
    t2 = t2_E.get().strip()
    d3 = d3_E.get().strip()


    # Check if required fields are empty
    if not all((a1, a2, a3, t1, t2, d3)):
        messagebox.showerror("Error", "Please fill all required fields.")
        return

    # Convert empty fields to '0'
    a1 = a1 if a1 else '0'
    a2 = a2 if a2 else '0'
    a3 = a3 if a3 else '0'
    t1 = t1 if t1 else '0'
    t2 = t2 if t2 else '0'
    d3 = d3 if d3 else '0'

    # Validate numeric input
    try:
        a1, a2, a3, t1, t2, d3 = [float(entry) for entry in (a1, a2, a3, t1, t2, d3)]
    except ValueError:
        messagebox.showerror("Error", "Invalid Input.")
        return

    a1 = float(a1_E.get())
    a2 = float(a2_E.get())
    a3 = float(a3_E.get())
    
    #link conversion to meters
    def cm_to_meter(a):
        m = 100 # 1 meter  = 100 cm
        return a/m

    a1 = cm_to_meter(a1)
    a2 = cm_to_meter(a2)
    a3 = cm_to_meter(a3)

    t1 = float(t1_E.get())
    t2 = float(t2_E.get())
    d3 = float(d3_E.get())

    t1 = (t1/180.0)*np.pi
    t2 = (t2/180.0)*np.pi
    d3 = cm_to_meter(d3)
    
    # Parametric Table

    PT = [[(0.0/180)*np.pi+t1, (90.0/180)*np.pi, 0, a1],
          [(90.0/180)*np.pi+t2, (90.0/180)*np.pi, 0, 0],
          [(0.0/180)*np.pi, (0.0/180)*np.pi, 0, a2+a3+d3]]

    # Homegeneous Transformation Matrix Formula

    i = 0
    H0_1 = [[np.cos(PT[i][0]), -np.sin(PT[i][0])*np.cos(PT[i][1]), np.sin(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.cos(PT[i][0])],
            [np.sin(PT[i][0]), np.cos(PT[i][0])*np.cos(PT[i][1]), -np.cos(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.sin(PT[i][0])],
            [0, np.sin(PT[i][1]), np.cos(PT[i][1]), PT[i][3]],
            [0, 0, 0, 1]]       

    i = 1
    H1_2 = [[np.cos(PT[i][0]), -np.sin(PT[i][0])*np.cos(PT[i][1]), np.sin(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.cos(PT[i][0])],
            [np.sin(PT[i][0]), np.cos(PT[i][0])*np.cos(PT[i][1]), -np.cos(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.sin(PT[i][0])],
            [0, np.sin(PT[i][1]), np.cos(PT[i][1]), PT[i][3]],
            [0, 0, 0, 1]] 

    i = 2
    H2_3 = [[np.cos(PT[i][0]), -np.sin(PT[i][0])*np.cos(PT[i][1]), np.sin(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.cos(PT[i][0])],
            [np.sin(PT[i][0]), np.cos(PT[i][0])*np.cos(PT[i][1]), -np.cos(PT[i][0])*np.sin(PT[i][1]), PT[i][2]*np.sin(PT[i][0])],
            [0, np.sin(PT[i][1]), np.cos(PT[i][1]), PT[i][3]],
            [0, 0, 0, 1]]


    #Multiply Matrices

    H0_2 = np.dot(H0_1, H1_2)
    H0_3 = np.dot(H0_2, H2_3)

    H0_1 = np.array(H0_1)

    X0_3 = H0_3[0,3]
    x_E.delete(0,END)
    x_E.insert(0,np.around(X0_3*100,3))

    Y0_3 = H0_3[1,3]
    y_E.delete(0,END)
    y_E.insert(0,np.around(Y0_3*100,3))

    Z0_3 = H0_3[2,3]
    z_E.delete(0,END)
    z_E.insert(0,np.around(Z0_3*100,3))

    save_to_excel_fk()

## Jacobian (Velocity Calculator)

    global J_sw    
    J_sw = Toplevel(window)
    J_sw.title("Velocity Calculator")
    J_sw.geometry("375x500")
    J_sw.configure(bg = "#FFFFFF")
    J_sw.resizable(False,False)
    J_sw.iconbitmap(r"assets\robotic-arm.ico")

        # Close existing J_sw window if it exists


    # Jacobian Matrix

    # 1. Linear/Translation Vectors
    Z_1 = np.array([[0], [0], [1]])  # The [0,0,1] vector

    #Row 1 to 3, Column 1
    J1a = [[1,0,0],
           [0,1,0],
           [0,0,1]]
    J1a = np.dot(J1a,Z_1)

    J1b_1 = H0_3[0:3,3:] #d0_3
    J1b_1 = np.array(J1b_1)

    J1b_2 = H0_1[0:3,3:] #d0_1
    J1b_2 = np.array(J1b_2)

    J1b= J1b_1 - J1b_2

    J1 = [[J1a[1,0]*J1b[2,0] - J1a[2,0]*J1b[1,0]],
        [J1a[2,0]*J1b[0,0] - J1a[0,0]*J1b[2,0]],
        [J1a[0,0]*J1b[1,0] - J1a[1,0]*J1b[0,0]]]

    J1 = np.array(J1)


    #Row 1 to 3, Column 2
    J2a = H0_1[0:3,0:3] #R0_1
    J2a = np.dot(J2a,Z_1)

    J2b_1 = H0_3[0:3,3:] #d0_3
    J2b_1 = np.array(J2b_1)

    J2b_2 = H0_1[0:3,3:] #d0_1
    J2b_2 = np.array(J2b_2)

    J2b= J2b_1 - J2b_2

    J2 = [[J2a[1,0]*J2b[2,0] - J2a[2,0]*J2b[1,0]],
        [J2a[2,0]*J2b[0,0] - J2a[0,0]*J2b[2,0]],
        [J2a[0,0]*J2b[1,0] - J2a[1,0]*J2b[0,0]]]

    J2 = np.array(J2)

    #Row 1 to 3, Column 3
    J3a = H0_2[0:3,0:3] 
    J3a = np.dot(J3a,Z_1)
    J3 = np.array(J3a)

    # 2. Rotation/Orientation Vectors
    #Row 4 to 6, Column 1
    J4 = J1a
    J4 = np.array(J4)
    print("J4 = ")
    print(J4)

    #Row 4 to 6, Column 2
    J5 = J2a
    J5 = np.array(J5)
    print("J5 = ")
    print(J5)

    J6 = np.array([[0], [0], [0]])

    # 3. Concatenated Jacobian Matrix
    JM1 = np.concatenate((J1, J2, J3), axis=1)
    JM2 = np.concatenate((J4, J5, J6), axis=1)
    J = np.concatenate((JM1, JM2), axis=0)

    print("Jacobian Matrix J = ")
    print(np.around(J, 3))

    # 4. Differential Equations
    T1_p, T2_p, d3_p = sp.symbols('θ1* θ2* d3*')

    q = [[T1_p], [T2_p], [d3_p]]

    E = np.dot(J,q)
    E = np.array(E)

    xp = E[0,0]
    yp = E[1,0]
    zp = E[2,0]
    ωx = E[3,0]
    ωy = E[4,0]
    ωz = E[5,0]

    def update_velo():
        T1p = T1_slider.get()
        T2p = T2_slider.get()
        d3p = d3_slider.get()

        q = np.array([[T1p],[T2p],[d3p]])
        E = np.dot(J,q)

        xp_e = E[0,0]
        entry_1.delete(0,END)
        entry_1.insert(0,str(xp_e))

        yp_e = E[1,0]
        entry_2.delete(0,END)
        entry_2.insert(0,str(yp_e))

        zp_e = E[2,0]
        entry_3.delete(0,END)
        entry_3.insert(0,str(zp_e))

        ωx_e = E[3,0]
        entry_4.delete(0,END)
        entry_4.insert(0,str(ωx_e))

        ωy_e = E[4,0]
        entry_5.delete(0,END)
        entry_5.insert(0,str(ωy_e))

        ωz_e = E[5,0]
        entry_6.delete(0,END)
        entry_6.insert(0,str(ωz_e))


    canvas1 = Canvas(
    J_sw,
    bg = "#FFFFFF",
    height = 500,
    width = 375,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
    )

    canvas1.place(x = 0, y = 0)
    canvas1.create_rectangle(
        0.0,
        0.0,
        376.0,
        500.0,
        fill="#262626",
        outline="")

    canvas1.create_text(
        210.0,
        39.0,
        anchor="nw",
        text="MATRIX",
        fill="#FFFFFF",
        font=("Metropolis Black", 32 * -1)
    )

    canvas1.create_text(
        38.0,
        39.0,
        anchor="nw",
        text="JACOBIAN",
        fill="#52FF00",
        font=("Metropolis Black", 32 * -1)
    )

    canvas1.create_text(
        20.0,
        70.0,
        anchor="nw",
        text="VELOCITY CALCULATOR OF SPHERICAL MANIPULATOR ",
        fill="#FFFFFF",
        font=("Metropolis Semi Bold", 12 * -1)
    )

    canvas1.create_text(
        61.0,
        113.0,
        anchor="nw",
        text="θ˙₁ = ",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        61.0,
        156.0,
        anchor="nw",
        text="θ˙₂ = ",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        61.0,
        198.0,
        anchor="nw",
        text="d˙₃ = ",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        61.0,
        238.0,
        anchor="nw",
        text="   ẋ =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        61.0,
        273.0,
        anchor="nw",
        text="   ẏ =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        62.0,
        308.0,
        anchor="nw",
        text="   ż =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        282.0,
        117.0,
        anchor="nw",
        text="rad/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        159.0,
        anchor="nw",
        text="rad/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        201.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        238.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        273.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        308.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    T1_slider = Scale(
        J_sw, 
        from_=0,to_=3.142,
        resolution= 0.01, 
        orient=HORIZONTAL,
        length=164,
        bg="#D9D9D9",
        bd=0, 
        relief="flat", 
        font=("Calibri", 10),
        )
    
    T2_slider = Scale(
        J_sw, 
        from_=0,to_=3.142,
        resolution= 0.01,  
        orient=HORIZONTAL,
        length=164, 
        bg="#D9D9D9",
        bd=0, 
        relief="flat", 
        font=("Calibri", 10)
        )
    
    d3_slider = Scale(
        J_sw, 
        from_=0,to_=5,
        resolution= 0.01,  
        orient=HORIZONTAL,
        length=164, 
        bg="#D9D9D9",
        bd=0, 
        relief="flat", 
        font=("Calibri", 10)
        )

    T1_slider.place(x=106.0, 
                    y=105.0)
    
    T2_slider.place(x=106.0, 
                    y=148.0)
    
    d3_slider.place(x=106.0, 
                    y=190.0)


    entry_image_1 = PhotoImage(
        file=relative_to_assets2("entry_1.png"))
    entry_bg_1 = canvas1.create_image(
        188.0,
        247.0,
        image=entry_image_1
    )
    entry_1 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_1.place(
        x=106.0,
        y=234.0,
        width=164.0,
        height=24.0
    )

    entry_image_2 = PhotoImage(
        file=relative_to_assets2("entry_2.png"))
    entry_bg_2 = canvas1.create_image(
        188.0,
        282.0,
        image=entry_image_2
    )
    entry_2 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_2.place(
        x=106.0,
        y=269.0,
        width=164.0,
        height=24.0
    )

    entry_image_3 = PhotoImage(
        file=relative_to_assets2("entry_3.png"))
    entry_bg_3 = canvas1.create_image(
        189.0,
        317.0,
        image=entry_image_3
    )
    entry_3 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_3.place(
        x=107.0,
        y=304.0,
        width=164.0,
        height=24.0
    )

    canvas1.create_text(
        61.0,
        343.0,
        anchor="nw",
        text="ωx =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        61.0,
        378.0,
        anchor="nw",
        text="ωy =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        62.0,
        413.0,
        anchor="nw",
        text="ωz =",
        fill="#FFFFFF",
        font=("Calibri", 18 * -1)
    )

    canvas1.create_text(
        282.0,
        343.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        378.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    canvas1.create_text(
        282.0,
        413.0,
        anchor="nw",
        text="cm/s",
        fill="#FFFFFF",
        font=("Calibri", 15 * -1)
    )

    entry_image_4 = PhotoImage(
        file=relative_to_assets2("entry_4.png"))
    entry_bg_4 = canvas1.create_image(
        188.0,
        352.0,
        image=entry_image_4
    )
    entry_4 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_4.place(
        x=106.0,
        y=339.0,
        width=164.0,
        height=24.0
    )

    entry_image_5 = PhotoImage(
        file=relative_to_assets2("entry_5.png"))
    entry_bg_5 = canvas1.create_image(
        188.0,
        387.0,
        image=entry_image_5
    )
    entry_5 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_5.place(
        x=106.0,
        y=374.0,
        width=164.0,
        height=24.0
    )

    entry_image_6 = PhotoImage(
        file=relative_to_assets2("entry_6.png"))
    entry_bg_6 = canvas1.create_image(
        189.0,
        422.0,
        image=entry_image_6
    )
    entry_6 = Entry(
        J_sw,
        bd=0,
        font=('Calibri', 13),
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0
    )
    entry_6.place(
        x=107.0,
        y=409.0,
        width=164.0,
        height=24.0
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets2("button_1.png"))
    button_1 = Button(
        J_sw,
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=update_velo,
        relief="flat"
    )
    button_1.place(
        x=129.0,
        y=452.0,
        width=117.42608642578125,
        height=30.26706314086914
    )


    #Create links
    #robot_variable = DHRobot([RevoluteDH(d,r,alpha,offset=theta,qlim)])
    #robot_variable = DHRobot([PrismaticDH(d=0,r,alpha,offset=d,qlim)])
    Spherical = DHRobot([
        RevoluteDH(a1,0,(90.0/180.0)*np.pi,(0.0/180.0)*np.pi,qlim=[-np.pi/2,np.pi/2]),
        RevoluteDH(0,0,(90.0/180.0)*np.pi,(90.0/180.0)*np.pi,qlim=[-np.pi/4,np.pi/4]),
        PrismaticDH(0,0,(0.0/180)*np.pi,a2+a3,qlim=[0,d3])
    ], name="Spherical")


    q0 = np.array([0,0,0]) # origin
    q1 = np.array([t1,t2,d3])
    traj1 = rtb.jtraj(q0,q1,30)

    x1 = -1.0
    x2 = 1.0
    y1 = -1.0
    y2 = 1.0
    z1 = 0
    z2 = 1.0

    # Plot commands
    Spherical.plot(traj1.q, limits=[x1,x2,y1,y2,z1,z2], block=True)


def i_k():
    # Get values from entry widgets
    x_pos = x_E.get().strip()
    y_pos = y_E.get().strip()
    z_pos = z_E.get().strip()
    a1 = a1_E.get().strip()
    a2 = a2_E.get().strip()
    a3 = a3_E.get().strip()

    # Check if required fields are empty
    if not all((a1, a2, a3, x_pos, y_pos, z_pos)):
        messagebox.showerror("Error", "Please fill all required fields.")
        return
    
    # Convert empty fields to '0'
    a1 = a1 if a1 else '0'
    a2 = a2 if a2 else '0'
    a3 = a3 if a3 else '0'
    x_pos = x_pos if x_pos else '0'
    y_pos = y_pos if y_pos else '0'
    z_pos = z_pos if z_pos else '0'

    # Validate numeric input
    try:
        a1, a2, a3, x_pos, y_pos, z_pos = [float(entry) for entry in (a1, a2, a3, x_pos, y_pos, z_pos)]
    except ValueError:
        messagebox.showerror("Error", "Invalid Input.")
        return

    # Convert values to float
    x_pos = float(x_E.get()) / 100
    y_pos = float(y_E.get()) / 100
    z_pos = float(z_E.get()) / 100
    a1 = float(a1_E.get()) / 100  # Convert a1 from cm to meters
    a2 = float(a2_E.get()) / 100  # Convert a2 from cm to meters
    a3 = float(a3_E.get()) / 100  # Convert a3 from cm to meters

    # Perform inverse kinematics calculations
    try:
        t1_ik = np.arctan(y_pos / x_pos)  # Formula 1
        r1 = np.sqrt((x_pos) ** 2 + (y_pos) ** 2)  # Formula 2
        r2 = z_pos - a1  # Formula 3
        t2_ik = np.arctan(r2 / r1)  # Formula 4
        d3_ik_cm = np.sqrt((r1 ** 2) + (r2 ** 2)) - a2 - a3  # Formula 5
    except ZeroDivisionError:
        messagebox.showerror("Error", "Undefined.")
        return

    # Update the entry widgets with the calculated values
    t1_E.delete(0, END)
    t1_E.insert(0, np.around(t1_ik * 180 / np.pi, 3))

    t2_E.delete(0, END)
    t2_E.insert(0, np.around(t2_ik * 180 / np.pi, 3))

    d3_E.delete(0, END)
    d3_E.insert(0, np.around(d3_ik_cm * 100, 3))  # Display in cm

    save_to_excel_ik()

    # Plot commands
    Spherical = DHRobot([
        RevoluteDH(a1,0,(90.0/180.0)*np.pi,(0.0/180.0)*np.pi,qlim=[-np.pi/2,np.pi/2]),
        RevoluteDH(0,0,(90.0/180.0)*np.pi,(90.0/180.0)*np.pi,qlim=[-np.pi/4,np.pi/4]),
        PrismaticDH(0,0,(0.0/180)*np.pi,a2+a3,qlim=[0,d3_ik_cm])
    ], name="Spherical")

    q0 = np.array([0,0,0]) # origin
    q1 = np.array([t1_ik, t2_ik, d3_ik_cm])  # Use d3_ik_m (in meters)
    
    traj1 = rtb.jtraj(q0,q1,30)

    x1 = -1.0
    x2 = 1.0
    y1 = -1.0
    y2 = 1.0
    z1 = 0
    z2 = 1.0

    # Plot commands
    Spherical.plot(traj1.q, limits=[x1,x2,y1,y2,z1,z2], block=True)



canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 600,
    width = 900,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    0.0,
    900.0,
    600.0,
    fill="#262626",
    outline="")

canvas.create_text(
    130.0,
    49.0,
    anchor="nw",
    text="SPHERICAL",
    fill="#51FF00",
    font=("Metropolis Black", 48 * -1)
)

canvas.create_text(
    407.0,
    49.0,
    anchor="nw",
    text="MANIPULATOR",
    fill="#FFFFFF",
    font=("Metropolis Black", 48 * -1)
)

canvas.create_text(
    253.0,
    92.0,
    anchor="nw",
    text="FORWARD AND INVERSE KINEMATICS CALCULATOR",
    fill="#FFFFFF",
    font=("Metropolis Black", 15 * -1)
)

canvas.create_text(
    238.0,
    147.0,
    anchor="nw",
    text="Link Lengths",
    fill="#FFFFFF",
    font=("Calibri Italic", 20 * -1)
)

canvas.create_text(
    110.0,
    367.0,
    anchor="nw",
    text="Joint Variables",
    fill="#FFFFFF",
    font=("Calibri Italic", 20 * -1)
)

canvas.create_text(
    335.0,
    367.0,
    anchor="nw",
    text="Position Vector",
    fill="#FFFFFF",
    font=("Calibri Italic", 20 * -1)
)

canvas.create_text(
    190.0,
    184.0,
    anchor="nw",
    text="a₁:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    359.0,
    184.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    73.0,
    403.0,
    anchor="nw",
    text="θ₁:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    242.0,
    403.0,
    anchor="nw",
    text="°",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    73.0,
    443.0,
    anchor="nw",
    text="θ₂:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    242.0,
    443.0,
    anchor="nw",
    text="°",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    73.0,
    483.0,
    anchor="nw",
    text="d₃:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    242.0,
    483.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    307.0,
    403.0,
    anchor="nw",
    text="x:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    469.0,
    403.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    307.0,
    443.0,
    anchor="nw",
    text="y:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    307.0,
    483.0,
    anchor="nw",
    text="z:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    471.0,
    443.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    471.0,
    483.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    190.0,
    224.0,
    anchor="nw",
    text="a₂:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    359.0,
    224.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    190.0,
    264.0,
    anchor="nw",
    text="a₃:",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

canvas.create_text(
    359.0,
    264.0,
    anchor="nw",
    text="cm",
    fill="#FFFFFF",
    font=("Calibri", 16 * -1)
)

entry_image_3 = PhotoImage(
    file=relative_to_assets1("entry_3.png"))
entry_bg_3 = canvas.create_image(
    284.5,
    194.5,
    image=entry_image_3
)
a1_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
a1_E.place(
    x=218.0,
    y=182.0,
    width=133.0,
    height=29.0
)

entry_image_2 = PhotoImage(
    file=relative_to_assets1("entry_2.png"))
entry_bg_2 = canvas.create_image(
    284.5,
    234.5,
    image=entry_image_2
)
a2_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
a2_E.place(
    x=218.0,
    y=222.0,
    width=133.0,
    height=29.0
)

entry_image_1 = PhotoImage(
    file=relative_to_assets1("entry_1.png"))
entry_bg_1 = canvas.create_image(
    284.5,
    274.5,
    image=entry_image_1
)
a3_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
a3_E.place(
    x=218.0,
    y=260.0,
    width=133.0,
    height=29.0
)

entry_image_4 = PhotoImage(
    file=relative_to_assets1("entry_4.png"))
entry_bg_4 = canvas.create_image(
    168.0,
    414.0,
    image=entry_image_4
)
t1_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
t1_E.place(
    x=102.0,
    y=398.0,
    width=132.0,
    height=30.0
)

entry_image_5 = PhotoImage(
    file=relative_to_assets1("entry_5.png"))
entry_bg_5 = canvas.create_image(
    168.0,
    453.5,
    image=entry_image_5
)
t2_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
t2_E.place(
    x=102.0,
    y=438.5,
    width=132.0,
    height=29.0
)

entry_image_6 = PhotoImage(
    file=relative_to_assets1("entry_6.png"))
entry_bg_6 = canvas.create_image(
    168.0,
    493.5,
    image=entry_image_6
)
d3_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
d3_E.place(
    x=102.0,
    y=478.5,
    width=132.0,
    height=29.0
)

entry_image_7 = PhotoImage(
    file=relative_to_assets1("entry_7.png"))
entry_bg_7 = canvas.create_image(
    395.0,
    414.0,
    image=entry_image_7
)
x_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
x_E.place(
    x=329.0,
    y=398.0,
    width=132.0,
    height=30.0
)

entry_image_8 = PhotoImage(
    file=relative_to_assets1("entry_8.png"))
entry_bg_8 = canvas.create_image(
    395.0,
    453.5,
    image=entry_image_8
)
y_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
y_E.place(
    x=329.0,
    y=438.5,
    width=132.0,
    height=29.0
)

entry_image_9 = PhotoImage(
    file=relative_to_assets1("entry_9.png"))
entry_bg_9 = canvas.create_image(
    395.0,
    495.0,
    image=entry_image_9
)
z_E = Entry(window,
    font=('Calibri', 13),
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
z_E.place(
    x=329.0,
    y=479.0,
    width=132.0,
    height=30.0
)

button_image_1 = PhotoImage(
    file=relative_to_assets1("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=f_k,
    relief="flat"
)
button_1.place(
    x=108.0,
    y=522.73291015625,
    width=118.0,
    height=30.26708984375
)

button_image_2 = PhotoImage(
    file=relative_to_assets1("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=i_k,
    relief="flat"
)
button_2.place(
    x=335.9855041503906,
    y=522.73291015625,
    width=117.42608642578125,
    height=30.26706314086914
)

button_image_3 = PhotoImage(
    file=relative_to_assets1("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=reset,
    relief="flat"
)
button_3.place(
    x=235.0,
    y=301.0,
    width=99.0,
    height=30.44510269165039
)

canvas.create_rectangle(
    530.0,
    131.0,
    828.0,
    347.0,
    fill="#262626",
    outline="")

image_image_1 = PhotoImage(
    file=relative_to_assets1("image_1.png"))
image_1 = canvas.create_image(
    679.0,
    239.0,
    image=image_image_1
)

canvas.create_rectangle(
    530.0,
    359.0,
    828.0,
    560.0,
    fill="#262626",
    outline="")

image_image_2 = PhotoImage(
    file=relative_to_assets1("image_2.png"))
image_2 = canvas.create_image(
    679.0,
    459.0,
    image=image_image_2
)

# Create the checkbutton
save_to_excel_checkbutton = Checkbutton(
    window,
    variable=save_to_excel_var,
    command=toggle_save_to_excel,
    bg="#262626",
    activebackground="#262626",
    selectcolor="#51FF00",
)
save_to_excel_checkbutton.place(
    x=100,
    y=565
)

save_to_excel_label = Label(
    text='Record Data',
    font= ('Metropolis Semi Bold', 10),
    fg= '#FFFFFF',
    bg= '#262626'
)

save_to_excel_label.place(
    x=120,
    y=569
)
view_excel_button = Button(
    text= "View Excel",
    fg= 'black',
    bg= '#51FF00',
    font= ('Metropolis Extra Bold', 10),
    command= view_excel_file
)

view_excel_button.place(
    x=10,
    y=567
)

canvas.create_text(
    675,
    578,
    text="Path & Trajectory:",
    fill="#FFFFFF",
    font=("Metropolis Semi bold", 10)
)

path_pick = Button(
    text= "Pick&Place",
    fg= 'black',
    bg= '#51FF00',
    font= ('Metropolis Extra Bold', 10),
    command= p_t_pick
)

path_pick.place(
    x=810,
    y=567
)

path_weld = Button(
    text= "Welding",
    fg= 'black',
    bg= '#51FF00',
    font= ('Metropolis Extra Bold', 10),
    command= p_t_pick
)

path_weld.place(
    x=740,
    y=567
)


window.resizable(False, False)
window.mainloop()